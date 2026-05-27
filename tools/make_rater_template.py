"""
tools/make_rater_template.py
----------------------------
Builds tools/results/rater_template.xlsx, an Excel sheet with:

  Sheet 1: Engine vs Human (60 rows of scenarios)
      For each scenario, three rater columns IN/OUT and one ENGINE column.
      Computes percent agreement and Cohen's kappa between every pair.

  Sheet 2: Hint Quality (20 rows of paired hints)
      For each scenario, three raters score hint_A and hint_B on 1-5.
      Computes mean rating per condition (after the blind is broken)
      and the difference.

Run from project root:
    python -m tools.make_rater_template
"""
from __future__ import annotations
import csv
import os
import sys

THIS_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(THIS_DIR)
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from tools.scenarios import ALL_SCENARIOS

RESULTS_DIR = os.path.join(THIS_DIR, "results")
os.makedirs(RESULTS_DIR, exist_ok=True)
OUT_PATH = os.path.join(RESULTS_DIR, "rater_template.xlsx")


HEADER_FILL = PatternFill("solid", fgColor="14213D")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
LABEL_FILL  = PatternFill("solid", fgColor="2E4057")
LABEL_FONT  = Font(bold=True, color="E0E0E0")
STAT_FILL   = PatternFill("solid", fgColor="FFD93D")
STAT_FONT   = Font(bold=True, color="14213D", size=12)
BORDER = Border(
    left=Side(style="thin", color="888888"),
    right=Side(style="thin", color="888888"),
    top=Side(style="thin", color="888888"),
    bottom=Side(style="thin", color="888888"),
)


def _style_header_row(ws, row, n_cols):
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 32


def _set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_engine_vs_human(wb):
    ws = wb.create_sheet("Engine vs Human")
    headers = [
        "Scenario ID", "Name", "Topology",
        "Engine verdict (IN/OUT)",
        "Rater 1 verdict", "Rater 2 verdict", "Rater 3 verdict",
        "All agree?",
    ]
    _set_widths(ws, [12, 36, 9, 16, 14, 14, 14, 11])
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, len(headers))

    # Engine verdicts are filled from the audit CSV if available
    audit_csv = os.path.join(RESULTS_DIR, "evaluation_results.csv")
    engine_verdicts = {}
    if os.path.exists(audit_csv):
        with open(audit_csv) as f:
            for row in csv.DictReader(f):
                engine_verdicts[row["id"]] = row["engine"]

    for i, s in enumerate(ALL_SCENARIOS, 2):
        ws.cell(row=i, column=1, value=s["id"])
        ws.cell(row=i, column=2, value=s["name"])
        ws.cell(row=i, column=3, value=s["topology"])
        ws.cell(row=i, column=4, value=engine_verdicts.get(s["id"], ""))
        # Rater columns left blank
        for col in (5, 6, 7):
            ws.cell(row=i, column=col, value="")
            ws.cell(row=i, column=col).alignment = Alignment(horizontal="center")
        # "All agree" formula
        f = (f'=IF(AND(D{i}<>"",E{i}<>"",F{i}<>"",G{i}<>"",'
             f'D{i}=E{i},E{i}=F{i},F{i}=G{i}),"Y","N")')
        ws.cell(row=i, column=8, value=f)
        for col in range(1, 9):
            ws.cell(row=i, column=col).border = BORDER

    # Stats block below the table
    last_data_row = 1 + len(ALL_SCENARIOS)
    stat_row = last_data_row + 3

    def label(row, col, txt):
        c = ws.cell(row=row, column=col, value=txt)
        c.fill = LABEL_FILL
        c.font = LABEL_FONT
        c.alignment = Alignment(horizontal="right")

    def stat(row, col, formula):
        c = ws.cell(row=row, column=col, value=formula)
        c.fill = STAT_FILL
        c.font = STAT_FONT
        c.alignment = Alignment(horizontal="center")
        c.number_format = "0.000"

    label(stat_row, 4, "Engine vs Rater 1 % agreement")
    stat(stat_row, 5,
         f'=SUMPRODUCT((D2:D{last_data_row}=E2:E{last_data_row})*'
         f'(D2:D{last_data_row}<>""))/'
         f'SUMPRODUCT((D2:D{last_data_row}<>"")*(E2:E{last_data_row}<>""))')

    label(stat_row + 1, 4, "Engine vs Rater 2 % agreement")
    stat(stat_row + 1, 5,
         f'=SUMPRODUCT((D2:D{last_data_row}=F2:F{last_data_row})*'
         f'(D2:D{last_data_row}<>""))/'
         f'SUMPRODUCT((D2:D{last_data_row}<>"")*(F2:F{last_data_row}<>""))')

    label(stat_row + 2, 4, "Engine vs Rater 3 % agreement")
    stat(stat_row + 2, 5,
         f'=SUMPRODUCT((D2:D{last_data_row}=G2:G{last_data_row})*'
         f'(D2:D{last_data_row}<>""))/'
         f'SUMPRODUCT((D2:D{last_data_row}<>"")*(G2:G{last_data_row}<>""))')

    label(stat_row + 4, 4, "Rater 1 vs Rater 2 % agreement")
    stat(stat_row + 4, 5,
         f'=SUMPRODUCT((E2:E{last_data_row}=F2:F{last_data_row})*'
         f'(E2:E{last_data_row}<>""))/'
         f'SUMPRODUCT((E2:E{last_data_row}<>"")*(F2:F{last_data_row}<>""))')

    # Cohen's kappa (Rater 1 vs Rater 2) - written out as formulae
    # kappa = (Po - Pe) / (1 - Pe)
    # where Po = % observed agreement, Pe = expected agreement by chance
    n_range = f"E2:E{last_data_row}"
    m_range = f"F2:F{last_data_row}"
    label(stat_row + 5, 4, "Cohen's kappa (R1 vs R2)")
    formula = (
        f'=LET('
          f'po,SUMPRODUCT(({n_range}={m_range})*({n_range}<>""))/'
            f'SUMPRODUCT(({n_range}<>"")*({m_range}<>"")),'
          f'p_in1,COUNTIF({n_range},"IN")/COUNTA({n_range}),'
          f'p_in2,COUNTIF({m_range},"IN")/COUNTA({m_range}),'
          f'p_out1,COUNTIF({n_range},"OUT")/COUNTA({n_range}),'
          f'p_out2,COUNTIF({m_range},"OUT")/COUNTA({m_range}),'
          f'pe,p_in1*p_in2+p_out1*p_out2,'
          f'(po-pe)/(1-pe))'
    )
    stat(stat_row + 5, 5, formula)

    # Notes
    label(stat_row + 7, 4, "Notes:")
    ws.cell(row=stat_row + 7, column=5,
            value="Enter IN or OUT in rater columns. Stats compute live.")
    ws.cell(row=stat_row + 7, column=5).alignment = Alignment(horizontal="left")


def build_hint_quality(wb):
    ws = wb.create_sheet("Hint Quality")
    headers = [
        "Scenario ID", "Topology", "MDP action",
        "Hint A", "Hint B",
        "R1 score A", "R1 score B",
        "R2 score A", "R2 score B",
        "R3 score A", "R3 score B",
        "A is MDP?",
        "MDP score (R1)", "MDP score (R2)", "MDP score (R3)",
        "Baseline score (R1)", "Baseline score (R2)", "Baseline score (R3)",
    ]
    _set_widths(ws, [12, 9, 18, 50, 50,
                     11, 11, 11, 11, 11, 11, 10,
                     14, 14, 14, 14, 14, 14])
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, len(headers))

    # Fill from existing hint_comparison.csv if available
    hint_csv = os.path.join(RESULTS_DIR, "hint_comparison.csv")
    if os.path.exists(hint_csv):
        with open(hint_csv) as f:
            data = list(csv.DictReader(f))
    else:
        data = []

    for i, row in enumerate(data, 2):
        ws.cell(row=i, column=1, value=row["scenario_id"])
        ws.cell(row=i, column=2, value=row["topology"])
        ws.cell(row=i, column=3, value=row["mdp_action"])
        ws.cell(row=i, column=4, value=row["hint_A"])
        ws.cell(row=i, column=5, value=row["hint_B"])
        ws.cell(row=i, column=12, value=row["_hidden_A_is_mdp"])
        for col in range(1, 19):
            ws.cell(row=i, column=col).border = BORDER
            ws.cell(row=i, column=col).alignment = Alignment(
                wrap_text=True, vertical="top")
        # Resolved-score columns (use IF to route hint_A/B to MDP or baseline)
        for r_idx, (score_a_col, score_b_col, mdp_out_col, base_out_col) in enumerate(
            [(6, 7, 13, 16), (8, 9, 14, 17), (10, 11, 15, 18)]
        ):
            sa = f"{get_column_letter(score_a_col)}{i}"
            sb = f"{get_column_letter(score_b_col)}{i}"
            ws.cell(row=i, column=mdp_out_col, value=f'=IF(L{i}=TRUE,{sa},{sb})')
            ws.cell(row=i, column=base_out_col, value=f'=IF(L{i}=TRUE,{sb},{sa})')

    last_row = 1 + len(data)
    stat_row = last_row + 3

    def label(row, col, txt):
        c = ws.cell(row=row, column=col, value=txt)
        c.fill = LABEL_FILL
        c.font = LABEL_FONT
        c.alignment = Alignment(horizontal="right")

    def stat(row, col, formula):
        c = ws.cell(row=row, column=col, value=formula)
        c.fill = STAT_FILL
        c.font = STAT_FONT
        c.alignment = Alignment(horizontal="center")
        c.number_format = "0.00"

    if data:
        label(stat_row, 12, "Mean MDP score (R1)")
        stat(stat_row, 13, f'=AVERAGE(M2:M{last_row})')
        label(stat_row, 14, "Mean Baseline (R1)")
        stat(stat_row, 15, f'=AVERAGE(P2:P{last_row})')
        label(stat_row + 1, 12, "Mean MDP score (R2)")
        stat(stat_row + 1, 13, f'=AVERAGE(N2:N{last_row})')
        label(stat_row + 1, 14, "Mean Baseline (R2)")
        stat(stat_row + 1, 15, f'=AVERAGE(Q2:Q{last_row})')
        label(stat_row + 2, 12, "Mean MDP score (R3)")
        stat(stat_row + 2, 13, f'=AVERAGE(O2:O{last_row})')
        label(stat_row + 2, 14, "Mean Baseline (R3)")
        stat(stat_row + 2, 15, f'=AVERAGE(R2:R{last_row})')
        label(stat_row + 4, 12, "Overall MDP mean")
        stat(stat_row + 4, 13, f'=AVERAGE(M2:O{last_row})')
        label(stat_row + 4, 14, "Overall Baseline mean")
        stat(stat_row + 4, 15, f'=AVERAGE(P2:R{last_row})')


def build_readme(wb):
    ws = wb.create_sheet("Read me first", 0)
    _set_widths(ws, [70])
    rows = [
        ("LOGIC ADVOCATE TUTOR -- INTER-RATER SPREADSHEET", True),
        ("", False),
        ("This workbook has two sheets.", False),
        ("", False),
        ("Sheet 1: Engine vs Human", True),
        ("  - 60 scenarios from the audit, one row each.", False),
        ("  - The Engine verdict column is pre-filled from the harness run.", False),
        ("  - Up to three raters fill in IN or OUT in their column.", False),
        ("  - The stats block below computes Engine-vs-Rater agreement and", False),
        ("    Cohen's kappa between Rater 1 and Rater 2.", False),
        ("", False),
        ("Sheet 2: Hint Quality", True),
        ("  - 20 paired hints (MDP vs Baseline), one row each.", False),
        ("  - The labels are blinded as Hint A and Hint B.", False),
        ("  - Raters score each hint 1 to 5 on actionability.", False),
        ("  - The 'A is MDP?' column is the blinding key. The resolved-score", False),
        ("    columns (MDP score, Baseline score) compute automatically.", False),
        ("  - Means in the stat block at the bottom show which approach won.", False),
        ("", False),
        ("Workflow", True),
        ("  1. Run the harness to generate the underlying CSVs.", False),
        ("  2. Run make_rater_template.py to build this workbook.", False),
        ("  3. Share with two classmates. They fill in their columns.", False),
        ("  4. Read the stats block on each sheet. Cite the numbers in Ch 5.", False),
    ]
    for i, (txt, bold) in enumerate(rows, 1):
        c = ws.cell(row=i, column=1, value=txt)
        if bold:
            c.font = Font(bold=True, size=12)


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop the default sheet
    build_readme(wb)
    build_engine_vs_human(wb)
    build_hint_quality(wb)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print(f"Size: {os.path.getsize(OUT_PATH)} bytes")


if __name__ == "__main__":
    main()
